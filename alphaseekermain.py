from openpyxl import Workbook
import time
import pandas as pd
import operator
from alphaseekerclass import *
book = Workbook()
sheet = book.active
SAVEBOOK = "idk.xlsx"
SPECIFICSTOCK = input("do you want to look at specific stocks (1) or a huge list of stocks (2): ")
if(SPECIFICSTOCK != "1"):
    debug = input("do you want to use debug mode (recommended for testing the program unless you want to wait hours) Y/N: ")

def chooseStock(row):
    symbol = input("enter a stock press (1) if completed ")
    if(symbol == "1"):
        return True, row
    else:
        fullInfo = Stock(symbol)
        chooseStockSingle, y_predict, currentReturn = fullInfo.get_predicted()
        if(isinstance(chooseStockSingle,bool) == False):
            excelWrite(symbol, float(chooseStockSingle),float(y_predict), float(currentReturn), row)
            return False, row
        else:
            return False, row - 1

def notddos(overHeat):
    if(overHeat >= 25):
        print("overheat check")
        time.sleep(-time.time()%180)
        return 0
    else:
        overHeat += 1
        return overHeat

def parserFunction(row,stockList):
    errorCount = 0
    overHeat = 0
    howmany = len(stockList)
    for symbol in stockList:
        print("-----------------on %s, %d to go------------------" %(symbol,howmany))
        fullInfo = Stock(symbol)
        #overHeat = notddos(overHeat)
        #try:
        chooseStockSingle, y_predict, currentReturn = fullInfo.get_predicted()
        errorCount = 0
        if(isinstance(chooseStockSingle,bool) == False):
            excelWrite(symbol, float(chooseStockSingle),float(y_predict), float(currentReturn), row)
            row += 1
        else:
            pass
        # except:
        #     print("--------------ERROR 404--------------------")
        #     errorCount += 1
        #     if(errorCount > 5):
        #         time.sleep(-time.time()%75)
        #         try:
        #             chooseStockSingle, y_predict, currentReturn = fullInfo.get_predicted()
        #             errorCount = 0
        #             if(isinstance(chooseStockSingle,bool) == False):
        #                 excelWrite(symbol, float(chooseStockSingle),float(y_predict), float(currentReturn), row)
        #                 row += 1
        #                 print(symbol, "good")
        #             else:
        #                 pass
        #         except:
        #             print(symbol, "pass")
        #             pass
        howmany = howmany - 1
def excelToTicker():
    workbook = pd.read_excel('nasdaq_screener_1619061711441.xlsx')
    #stockListDirty = workbook["Symbol"].values
    stockList = [x.replace("^","-") for x in workbook["Symbol"].values if isinstance(x,bool) == False]
    return stockList

def excelWrite(symbol, getPred,y_predict,currentReturn, row):
    sheet.cell(row, 1).value = symbol
    sheet.cell(row,2).value = getPred
    sheet.cell(row,3).value = y_predict
    sheet.cell(row,4).value = currentReturn


def writeExcelAxis():
    sheet.cell(1,2).value = "Percent return"
    sheet.cell(1,1).value = "Symbol"
    sheet.cell(1,3).value = "y_predict"
    sheet.cell(1,4).value = "Current Price"

def orderString(n):
    return str(n)+("th" if 4<=n%100<=20 else {1:"st",2:"nd",3:"rd"}.get(n%10, "th"))

def analyzeExcel():
    workbookAnalyze = pd.read_excel(SAVEBOOK)
    gainAnalyzer = workbookAnalyze["Percent return"].values
    symbolAnalyzer = workbookAnalyze["Symbol"].values
    dictionary = {}
    countSymbol = 0
    for symbol in symbolAnalyzer:
        dictionary[symbol] = gainAnalyzer[countSymbol]
        countSymbol += 1
    sorted_Dictionary = sorted(dictionary.items(),key=lambda x: x[1], reverse=True)
    count = 1
    for i in sorted_Dictionary:
        print("the %s gain is %s" %(orderString(count),i))
        if(count > 20):
            break
        else:
            count += 1



def main():
    writeExcelAxis()
    row = 2
    if(SPECIFICSTOCK == "1"):
        correct = False
        while not correct:
            correct, row = chooseStock(row)
            row += 1
    else:
        if(debug == "N"):
            stockList = excelToTicker()
        else:
            stockList = ["GME","ILMN","AAPL","RBLX"]
        parserFunction(2,stockList)
    book.save(SAVEBOOK)
    analyzeExcel()

main()